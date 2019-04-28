# -*- coding: utf-8 -*-
import openpyxl


def createXLSX():
    # 生成一个实例化的workbook对象 注意Workbook中首字母大写
    wb = openpyxl.Workbook()
    # 获取当前工作表 sheet
    ws = wb.active
    # 修改工作表的名称
    ws.title = "tmp_sheet"
    # 向单个单元格传入数据
    ws['A1'] = 'name'
    ws['B1'] = 'age'
    data = {
        '老王': 50,
        '老李': 21,
        '老张': 32,
    }
    data_excel = []
    # 将字典中的每对数据（键，值）以列表形式传入data_excel列表
    # [['老李', 21], ['老王', 50], ['老张', 32]]
    for each in data:
        data_excel.append([each, data[each]])

    # 将列表里的数据写入数据表中
    for each in data_excel:
        ws.append(each)

    # 使用 Workbook.create_sheet() 创建新的worksheets,且定义表在工作簿的位置
    ws1 = wb.create_sheet('Mysheetfirst', 0)
    ws1['A1'] = "test"

    # 保存文件
    wb.save("./test.xlsx")

def opXLSX():
    wb = openpyxl.load_workbook(r"./test.xlsx")
    wb.create_sheet("new sheet")
    #打印文件中所有的表名
    print(wb.sheetnames)
    #删除表格 参数不能是字符串名称，而是一个工作表对象
    wb.remove(wb['new sheet']) #等价del wb['new_ws2']
    print(wb.sheetnames)
    #访问单元格
    ws = wb['tmp_sheet']

    print(ws['A1'])
    #打印单元格A1的值
    print(ws['A1'].value)
    #每个单元格对象拥有row(行), column(列), coordinate(坐标)
    a1 = ws['A1']
    ws['A1'] = 'china'
    print(a1.row,a1.column,a1.coordinate)
    #offset()方法偏移单元格， 第一个参数指定行， 第二个参数指定列
    a2 = a1.offset(1,0)
    print(a2.value)
    b1 = a1.offset(0,1)
    print(b1.value)
    #列数从字母A开始 有超过个位的列数
    # 列数转化为字母
    print(openpyxl.cell.cell.get_column_letter(666))
    # 字母转化为列数
    print(openpyxl.cell.cell.column_index_from_string('ZH'))
    #用切片访问数据
    for each_cell in ws['A1':'B4']:
        #print(each_cell)
        for each in each_cell:
            print(each.coordinate,end = ":")
            print(each.value,end ="  ")
        print("\n",end='')
    #按行打印
    for each_row in ws.rows:
        print(each_row)
    #按列打印
    for each_column in ws.columns:
        print(each_column)
    #打印第2列
    for each_row in ws.rows:
        print(each_row[1].value)
    #制定范围访问 C1：E3
    for each_row in ws.iter_rows(min_row=1, min_col=3, max_row=3, max_col=5):
        print(each_row)
    #表拷贝
    ws1 = wb.copy_worksheet(ws)
    wb.save("./test.xlsx")

def add():
    wb = openpyxl.load_workbook(r"./test.xlsx" )
    ws = wb['tmp_sheet']
    ws['A1'] = 'name'
    ws['B1'] = 'age'
    data = {
        'Mary': 28,
        'Jim': 74,
        'Suny': 25,
    }
    data_excel = []
    for each in data:
        data_excel.append([each, data[each]])

    # 将列表里的数据写入数据表中
    for each in data_excel:
        print(each)
        ws.append(each)
    wb.save("./test.xlsx")

def main():
    #createXLSX()
    #opXLSX()
    add()


if __name__ == "__main__":
    main()
